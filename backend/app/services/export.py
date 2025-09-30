from io import BytesIO
from datetime import datetime
from typing import Any, Dict, List, Optional
import csv
import json

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sqlalchemy.orm import Session
from app.models.iht import IHTProfile, Gift, Trust, Asset
from app.models.financial import BalanceSheet, ProfitLoss, CashFlow
from app.models.product import Product, Pension, Investment, Protection


class ExportService:
    """Service for exporting data in various formats"""

    @staticmethod
    def generate_iht_pdf(iht_calculation: Dict[str, Any]) -> BytesIO:
        """Generate PDF report for IHT calculation"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                                 topMargin=72, bottomMargin=18)

        # Container for the 'Flowable' objects
        elements = []

        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            textColor=colors.HexColor('#1a73e8'),
            spaceAfter=30
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a73e8'),
            spaceAfter=12
        )

        # Title
        elements.append(Paragraph("UK Inheritance Tax Report", title_style))
        elements.append(Spacer(1, 12))

        # Date
        date_str = datetime.now().strftime("%B %d, %Y")
        elements.append(Paragraph(f"Generated on: {date_str}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Estate Summary
        elements.append(Paragraph("Estate Summary", heading_style))

        summary_data = [
            ['Description', 'Amount (£)'],
            ['Total Estate Value', f"£{iht_calculation.get('total_assets', 0):,.2f}"],
            ['Total Liabilities', f"£{iht_calculation.get('total_liabilities', 0):,.2f}"],
            ['Net Estate Value', f"£{iht_calculation.get('net_estate', 0):,.2f}"],
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Tax Calculation
        elements.append(Paragraph("Tax Calculation", heading_style))

        tax_data = [
            ['Description', 'Amount (£)'],
            ['Taxable Estate', f"£{iht_calculation.get('taxable_estate', 0):,.2f}"],
            ['Standard Nil-Rate Band', f"£{iht_calculation.get('nil_rate_band', 325000):,.2f}"],
            ['Residence Nil-Rate Band', f"£{iht_calculation.get('residence_nil_rate_band', 0):,.2f}"],
            ['Tax Rate', '40%'],
            ['Total Tax Due', f"£{iht_calculation.get('tax_due', 0):,.2f}"],
            ['Effective Tax Rate', f"{iht_calculation.get('effective_rate', 0):.2f}%"],
        ]

        tax_table = Table(tax_data, colWidths=[3*inch, 2*inch])
        tax_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -2), (-1, -2), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('BACKGROUND', (0, -2), (-1, -2), colors.HexColor('#ffe6e6')),
        ]))
        elements.append(tax_table)
        elements.append(Spacer(1, 20))

        # Assets Breakdown
        if 'assets' in iht_calculation and iht_calculation['assets']:
            elements.append(Paragraph("Assets Breakdown", heading_style))

            asset_data = [['Asset Type', 'Value (£)']]
            for asset in iht_calculation['assets']:
                asset_data.append([
                    asset.get('asset_type', 'Unknown'),
                    f"£{asset.get('value', 0):,.2f}"
                ])

            asset_table = Table(asset_data, colWidths=[3*inch, 2*inch])
            asset_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(asset_table)
            elements.append(Spacer(1, 20))

        # Gifts
        if 'gifts' in iht_calculation and iht_calculation['gifts']:
            elements.append(Paragraph("Gifts History", heading_style))

            gift_data = [['Date', 'Recipient', 'Amount (£)', 'Taper Relief']]
            for gift in iht_calculation['gifts']:
                gift_data.append([
                    gift.get('date', ''),
                    gift.get('recipient', ''),
                    f"£{gift.get('amount', 0):,.2f}",
                    f"{gift.get('taper_relief', 0):.0f}%"
                ])

            gift_table = Table(gift_data, colWidths=[1.5*inch, 2*inch, 1.25*inch, 1.25*inch])
            gift_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (3, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(gift_table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_financial_excel(
        balance_sheets: List[BalanceSheet],
        profit_losses: List[ProfitLoss],
        cash_flows: List[CashFlow]
    ) -> BytesIO:
        """Generate Excel file with financial statements"""
        buffer = BytesIO()
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        currency_format = '#,##0.00'

        # Balance Sheet
        if balance_sheets:
            ws_bs = wb.create_sheet("Balance Sheet")

            # Headers
            headers = ['Date', 'Current Assets', 'Fixed Assets', 'Total Assets',
                      'Current Liabilities', 'Long Term Liabilities', 'Total Liabilities',
                      'Equity', 'Total Liabilities & Equity']

            for col, header in enumerate(headers, 1):
                cell = ws_bs.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Data
            for row, bs in enumerate(balance_sheets, 2):
                ws_bs.cell(row=row, column=1, value=bs.period.strftime('%Y-%m-%d'))
                ws_bs.cell(row=row, column=2, value=bs.current_assets).number_format = currency_format
                ws_bs.cell(row=row, column=3, value=bs.fixed_assets).number_format = currency_format
                ws_bs.cell(row=row, column=4, value=bs.total_assets).number_format = currency_format
                ws_bs.cell(row=row, column=5, value=bs.current_liabilities).number_format = currency_format
                ws_bs.cell(row=row, column=6, value=bs.long_term_liabilities).number_format = currency_format
                ws_bs.cell(row=row, column=7, value=bs.total_liabilities).number_format = currency_format
                ws_bs.cell(row=row, column=8, value=bs.equity).number_format = currency_format
                ws_bs.cell(row=row, column=9, value=bs.total_liabilities + bs.equity).number_format = currency_format

            # Auto-adjust column widths
            for column_cells in ws_bs.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws_bs.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 30)

        # Profit & Loss
        if profit_losses:
            ws_pl = wb.create_sheet("Profit & Loss")

            # Headers
            headers = ['Date', 'Revenue', 'Cost of Sales', 'Gross Profit',
                      'Operating Expenses', 'Operating Profit', 'Interest', 'Tax',
                      'Net Profit', 'Gross Margin %', 'Net Margin %']

            for col, header in enumerate(headers, 1):
                cell = ws_pl.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Data
            for row, pl in enumerate(profit_losses, 2):
                ws_pl.cell(row=row, column=1, value=pl.period.strftime('%Y-%m-%d'))
                ws_pl.cell(row=row, column=2, value=pl.revenue).number_format = currency_format
                ws_pl.cell(row=row, column=3, value=pl.cost_of_sales).number_format = currency_format
                ws_pl.cell(row=row, column=4, value=pl.gross_profit).number_format = currency_format
                ws_pl.cell(row=row, column=5, value=pl.operating_expenses).number_format = currency_format
                ws_pl.cell(row=row, column=6, value=pl.operating_profit).number_format = currency_format
                ws_pl.cell(row=row, column=7, value=pl.interest).number_format = currency_format
                ws_pl.cell(row=row, column=8, value=pl.tax).number_format = currency_format
                ws_pl.cell(row=row, column=9, value=pl.net_profit).number_format = currency_format
                ws_pl.cell(row=row, column=10, value=pl.gross_margin).number_format = '0.00%'
                ws_pl.cell(row=row, column=11, value=pl.net_margin).number_format = '0.00%'

            # Auto-adjust column widths
            for column_cells in ws_pl.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws_pl.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 30)

        # Cash Flow
        if cash_flows:
            ws_cf = wb.create_sheet("Cash Flow")

            # Headers
            headers = ['Date', 'Operating Activities', 'Investing Activities',
                      'Financing Activities', 'Net Cash Flow', 'Opening Balance',
                      'Closing Balance']

            for col, header in enumerate(headers, 1):
                cell = ws_cf.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Data
            for row, cf in enumerate(cash_flows, 2):
                ws_cf.cell(row=row, column=1, value=cf.period.strftime('%Y-%m-%d'))
                ws_cf.cell(row=row, column=2, value=cf.operating_activities).number_format = currency_format
                ws_cf.cell(row=row, column=3, value=cf.investing_activities).number_format = currency_format
                ws_cf.cell(row=row, column=4, value=cf.financing_activities).number_format = currency_format
                ws_cf.cell(row=row, column=5, value=cf.net_cash_flow).number_format = currency_format
                ws_cf.cell(row=row, column=6, value=cf.opening_balance).number_format = currency_format
                ws_cf.cell(row=row, column=7, value=cf.closing_balance).number_format = currency_format

            # Auto-adjust column widths
            for column_cells in ws_cf.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws_cf.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 30)

        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], headers: Optional[List[str]] = None) -> BytesIO:
        """Export data to CSV format"""
        buffer = BytesIO()

        if not data:
            return buffer

        # Get headers from first item if not provided
        if not headers:
            headers = list(data[0].keys())

        # Write CSV using StringIO then encode to bytes
        from io import StringIO
        string_buffer = StringIO()
        writer = csv.DictWriter(string_buffer, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)

        # Convert to bytes
        buffer.write(string_buffer.getvalue().encode('utf-8'))
        buffer.seek(0)
        return buffer

    @staticmethod
    def import_from_csv(file_content: bytes, model_class: Any) -> List[Dict[str, Any]]:
        """Import data from CSV file"""
        from io import StringIO

        # Decode bytes to string
        string_content = file_content.decode('utf-8')
        string_buffer = StringIO(string_content)

        reader = csv.DictReader(string_buffer)
        data = []

        for row in reader:
            # Convert string values to appropriate types based on model
            processed_row = {}
            for key, value in row.items():
                if value:
                    # Try to convert to float for numeric fields
                    if any(keyword in key.lower() for keyword in ['amount', 'value', 'price', 'rate', 'balance']):
                        try:
                            processed_row[key] = float(value)
                        except ValueError:
                            processed_row[key] = value
                    # Try to convert to date for date fields
                    elif any(keyword in key.lower() for keyword in ['date', 'period']):
                        try:
                            processed_row[key] = datetime.strptime(value, '%Y-%m-%d').date()
                        except ValueError:
                            processed_row[key] = value
                    else:
                        processed_row[key] = value
                else:
                    processed_row[key] = None

            data.append(processed_row)

        return data

    @staticmethod
    def create_backup(db: Session) -> BytesIO:
        """Create a backup of all user data in JSON format"""
        backup_data = {
            'timestamp': datetime.now().isoformat(),
            'version': '1.0',
            'data': {}
        }

        # Export IHT data
        iht_profiles = db.query(IHTProfile).all()
        if iht_profiles:
            backup_data['data']['iht_profiles'] = [
                {
                    'id': p.id,
                    'user_id': p.user_id,
                    'nil_rate_band': p.nil_rate_band,
                    'residence_nil_rate_band': p.residence_nil_rate_band,
                    'created_at': p.created_at.isoformat(),
                    'updated_at': p.updated_at.isoformat()
                }
                for p in iht_profiles
            ]

        gifts = db.query(Gift).all()
        if gifts:
            backup_data['data']['gifts'] = [
                {
                    'id': g.id,
                    'profile_id': g.profile_id,
                    'amount': g.amount,
                    'recipient': g.recipient,
                    'date': g.date.isoformat(),
                    'gift_type': g.gift_type,
                    'taper_relief': g.taper_relief
                }
                for g in gifts
            ]

        # Export Financial data
        balance_sheets = db.query(BalanceSheet).all()
        if balance_sheets:
            backup_data['data']['balance_sheets'] = [
                {
                    'id': bs.id,
                    'user_id': bs.user_id,
                    'period': bs.period.isoformat(),
                    'current_assets': bs.current_assets,
                    'fixed_assets': bs.fixed_assets,
                    'total_assets': bs.total_assets,
                    'current_liabilities': bs.current_liabilities,
                    'long_term_liabilities': bs.long_term_liabilities,
                    'total_liabilities': bs.total_liabilities,
                    'equity': bs.equity,
                    'created_at': bs.created_at.isoformat()
                }
                for bs in balance_sheets
            ]

        profit_losses = db.query(ProfitLoss).all()
        if profit_losses:
            backup_data['data']['profit_losses'] = [
                {
                    'id': pl.id,
                    'user_id': pl.user_id,
                    'period': pl.period.isoformat(),
                    'revenue': pl.revenue,
                    'cost_of_sales': pl.cost_of_sales,
                    'gross_profit': pl.gross_profit,
                    'operating_expenses': pl.operating_expenses,
                    'operating_profit': pl.operating_profit,
                    'interest': pl.interest,
                    'tax': pl.tax,
                    'net_profit': pl.net_profit,
                    'gross_margin': pl.gross_margin,
                    'net_margin': pl.net_margin,
                    'created_at': pl.created_at.isoformat()
                }
                for pl in profit_losses
            ]

        # Export Products
        products = db.query(Product).all()
        if products:
            backup_data['data']['products'] = [
                {
                    'id': p.id,
                    'user_id': p.user_id,
                    'product_type': p.product_type,
                    'provider': p.provider,
                    'product_name': p.product_name,
                    'value': p.value,
                    'created_at': p.created_at.isoformat()
                }
                for p in products
            ]

        # Convert to JSON bytes
        buffer = BytesIO()
        json_str = json.dumps(backup_data, indent=2)
        buffer.write(json_str.encode('utf-8'))
        buffer.seek(0)
        return buffer

    @staticmethod
    def restore_backup(db: Session, backup_content: bytes) -> Dict[str, Any]:
        """Restore user data from a backup file"""
        try:
            # Parse JSON backup
            backup_data = json.loads(backup_content.decode('utf-8'))

            if 'version' not in backup_data or 'data' not in backup_data:
                return {'success': False, 'error': 'Invalid backup format'}

            restored = {
                'iht_profiles': 0,
                'gifts': 0,
                'balance_sheets': 0,
                'profit_losses': 0,
                'products': 0
            }

            data = backup_data['data']

            # Restore IHT profiles
            if 'iht_profiles' in data:
                for profile_data in data['iht_profiles']:
                    profile = IHTProfile(
                        user_id=profile_data['user_id'],
                        nil_rate_band=profile_data['nil_rate_band'],
                        residence_nil_rate_band=profile_data['residence_nil_rate_band']
                    )
                    db.add(profile)
                    restored['iht_profiles'] += 1

            # Restore gifts
            if 'gifts' in data:
                for gift_data in data['gifts']:
                    gift = Gift(
                        profile_id=gift_data['profile_id'],
                        amount=gift_data['amount'],
                        recipient=gift_data['recipient'],
                        date=datetime.fromisoformat(gift_data['date']).date(),
                        gift_type=gift_data['gift_type'],
                        taper_relief=gift_data['taper_relief']
                    )
                    db.add(gift)
                    restored['gifts'] += 1

            # Restore balance sheets
            if 'balance_sheets' in data:
                for bs_data in data['balance_sheets']:
                    bs = BalanceSheet(
                        user_id=bs_data['user_id'],
                        period=datetime.fromisoformat(bs_data['period']).date(),
                        current_assets=bs_data['current_assets'],
                        fixed_assets=bs_data['fixed_assets'],
                        total_assets=bs_data['total_assets'],
                        current_liabilities=bs_data['current_liabilities'],
                        long_term_liabilities=bs_data['long_term_liabilities'],
                        total_liabilities=bs_data['total_liabilities'],
                        equity=bs_data['equity']
                    )
                    db.add(bs)
                    restored['balance_sheets'] += 1

            # Restore profit & loss
            if 'profit_losses' in data:
                for pl_data in data['profit_losses']:
                    pl = ProfitLoss(
                        user_id=pl_data['user_id'],
                        period=datetime.fromisoformat(pl_data['period']).date(),
                        revenue=pl_data['revenue'],
                        cost_of_sales=pl_data['cost_of_sales'],
                        gross_profit=pl_data['gross_profit'],
                        operating_expenses=pl_data['operating_expenses'],
                        operating_profit=pl_data['operating_profit'],
                        interest=pl_data['interest'],
                        tax=pl_data['tax'],
                        net_profit=pl_data['net_profit'],
                        gross_margin=pl_data['gross_margin'],
                        net_margin=pl_data['net_margin']
                    )
                    db.add(pl)
                    restored['profit_losses'] += 1

            # Restore products
            if 'products' in data:
                for product_data in data['products']:
                    product = Product(
                        user_id=product_data['user_id'],
                        product_type=product_data['product_type'],
                        provider=product_data['provider'],
                        product_name=product_data['product_name'],
                        value=product_data['value']
                    )
                    db.add(product)
                    restored['products'] += 1

            db.commit()

            return {
                'success': True,
                'restored': restored,
                'backup_date': backup_data['timestamp']
            }

        except Exception as e:
            db.rollback()
            return {'success': False, 'error': str(e)}